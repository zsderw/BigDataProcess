#!/usr/bin/python3

import openpyxl

wb = openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']

total = []
row_id = 1;
for row in ws:
    if row_id != 1:
        sum = ws.cell(row = row_id, column = 3).value * 0.3
        sum += ws.cell(row = row_id, column = 4).value * 0.35
        sum += ws.cell(row = row_id, column = 5).value * 0.34
        sum += ws.cell(row = row_id, column = 6).value
        ws.cell(row = row_id, column = 7).value = sum
        total.append(sum)
    row_id += 1

row_id = 1;
result = []
for i in total:
    rank = 1
    for j in total:
        if i < j:
            rank += 1
    result.append(rank)

row_id = 2;
sorting = len(result) / 10
for i in result :
    if i <= (sorting * 3 / 2):
        ws.cell(row = row_id, column = 8).value = 'A+'
    elif i <= (sorting * 3):
        ws.cell(row = row_id, column = 8).value = 'A0'
    elif i <= (sorting * 3 + (sorting * 7 - sorting * 3) / 2):
        ws.cell(row = row_id, column = 8).value = 'B+'
    elif i <= (sorting * 7):
        ws.cell(row = row_id, column = 8).value = 'B0'
    elif i <= (sorting * 7 + (sorting * 10 - sorting * 7) / 2):
        ws.cell(row = row_id, column = 8).value = 'C+'
    else:
        ws.cell(row = row_id, column = 8).value = 'C0'
    row_id += 1

wb.save("student.xlsx")
